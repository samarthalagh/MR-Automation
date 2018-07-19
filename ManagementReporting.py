#region Import section
import time
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import glob
import os
import openpyxl as xl
from shutil import copyfile
from openpyxl.utils import coordinate_from_string, column_index_from_string
from getpass import getpass
from selenium.common.exceptions import NoSuchElementException
#endregion

def getLatestPLfilepath(switch):
    if switch == 'e':
        list_of_files = glob.iglob('C:\\Users\salagh\\Downloads\\Profit and Loss Transaction Details*.xlsx')
    else: 
        list_of_files = glob.iglob('C:\\Users\salagh\\Downloads\\Timecard Detail*.xlsx')
    latest_file = max(list_of_files, key=os.path.getctime)
    return latest_file
def IsLoginSuccessful():
    username = input("Username: ")
    password = getpass("Password: ")
    
    #Login
    driver.find_element_by_id("sawlogonuser").send_keys(username)
    driver.find_element_by_id("sawlogonpwd").send_keys(password)
    driver.find_element_by_id("sawlogonpwd").send_keys(Keys.RETURN)
    driver.implicitly_wait(30)
    
    try:
        driver.find_element_by_id("sawlogonuser")
    except NoSuchElementException:
        return True
    return False
# Copy worksheet data to pivot sheet
def MoveExpensesData():
    print('Starting copy operation for Expenses')
    
    src = getLatestPLfilepath('e')
    print('Source: ', src)
    pivotsheet = 'D:\\Box\\Box Sync\\Helios\\Helios Projects Repository\\189630 - Renault- Tech Initiatives\\Project Management\\Financials\\Profit and Loss Transaction Details.xlsx'
    print('Destination: ', pivotsheet)
    
    wb1 = xl.load_workbook(filename=src)
    ws1 = wb1.worksheets[0]
    print('Source Sheet Name: ', ws1.title)

    wb2 = xl.load_workbook(filename=pivotsheet)
    
    if 'Expenses' in wb2.sheetnames:
        del wb2['Expenses']
        print('Data sheet removed from destination')
    
    ws2 = wb2.create_sheet('Expenses')
    print('New Data sheet created in destination. Starting copy operation')
    
    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(pivotsheet)
    print('Expenses - Copy complete')

def MoveTimesheetData():
    print('Starting copy operation for Timesheets')
    src = getLatestPLfilepath('')
    print('Source: ', src)
    
    # Rename pivot sheet
    pivotsheet = 'D:\\Box\\Box Sync\\Helios\\Helios Projects Repository\\189630 - Renault- Tech Initiatives\\Project Management\\Financials\\Profit and Loss Transaction Details.xlsx'
    print('Destination: ', pivotsheet)
    
    #Get Source sheet name
    wb1 = xl.load_workbook(filename=src)
    ws1 = wb1.worksheets[0]
    print('Source Sheet Name: ', ws1.title)

    #Check if data already exists in destination
    #Delete if existing
    wb2 = xl.load_workbook(filename=pivotsheet)
    
    if 'Timecard Detail' in wb2.sheetnames:
        del wb2['Timecard Detail']
        print('Data sheet removed from destination')
    
    #Create new data sheet for Timecards
    ws2 = wb2.create_sheet('Timecard Detail')
    print('New Data sheet created in destination. Starting copy operation')
    
    #Copy all cells to destination
    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value
            
    #Correct AH4 (Total Hours - Actual) row number
    ws2.cell(4,34).value = ws2.cell(3,34).value
    ws2.cell(3,34).value = ''

    # xy = coordinate_from_string('A4') # returns ('A',4)
    #         col = column_index_from_string(xy[0]) # returns 1
    #         row = xy[1]
    #         #if row==
            
    wb2.save(pivotsheet)
    print('Timesheets - Copy complete')

def DownloadExcels():
    #region Open Chrome and Login    
    #driver = webdriver.Chrome("C:\\Softwares\\chromedriver_win32\\chromedriver.exe")
    driver.get("http://coloeabi03.sapient.com:9704/analytics/saw.dll?dashboard&PortalPath=%2Fusers%2Fsamalagh%2F_portal")
    driver.maximize_window()
    driver.implicitly_wait(20)
    #endregion
    
    #region Login
    loggedIn = False
    while not loggedIn:
        loggedIn = IsLoginSuccessful()
        print("User authenticated: ", loggedIn)
    #endregion

    #region Export expenses - P&L report
    print("Ready to extract expenses")
    driver.implicitly_wait(300)
    driver.find_elements_by_class_name("TapeDeckImageEna")[1].click()
    print('Identified element')
    time.sleep(30)
    driver.execute_script(
        "return saw.dashboard.exportToExcel(\'/shared/Financials/_portal/Profit and Loss - Preferred Currency\', \'PL2005 - Profit and Loss Detail\', true);")
    print('Saving File')
    time.sleep(5)
    #endregion
    
    #region Export Timetracking Report
    driver.get(
        "http://coloeabi03.sapient.com:9704/analytics/saw.dll?dashboard&PortalPath=%2Fshared%2FPeople%20Management%2F_portal%2FUtilization%20and%20Time%20Tracking")

    driver.maximize_window()
    driver.implicitly_wait(60)

    driver.find_elements_by_class_name("TapeDeckImageEna")[1].click()
    print('Identified element')
    time.sleep(60)
    driver.execute_script(
        "return saw.dashboard.exportToExcel(\'/shared/People Management/_portal/Utilization and Time Tracking\', \'Timecard Detail\', true);")
    time.sleep(5)
    print('Saving File')
    time.sleep(60)
    #endregion
    
    driver.quit()
    
# Download excel files from MR
#driver = webdriver.Chrome("C:\\Softwares\\chromedriver_win32\\chromedriver.exe")
#DownloadExcels()

# Update expenses in pivot sheet
#MoveExpensesData()

# Update timesheets in pivot sheet
MoveTimesheetData()

# Update data sources on pivot sheets.

# Upload in Box
