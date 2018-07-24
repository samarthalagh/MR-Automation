
#region Import section
import time
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import glob
import os
import openpyxl as xl
from shutil import copyfile
from openpyxl.utils import coordinate_from_string, column_index_from_string
from getpass import getpass
from selenium.common.exceptions import NoSuchElementException
#endregion

def getLatestPLfilepath(switch):
    if switch == 'e':
        list_of_files = glob.iglob('C:\\Users\salagh\\Downloads\\Profit and Loss Transaction Details*.xlsx')
    else: 
        list_of_files = glob.iglob('C:\\Users\salagh\\Downloads\\Timecard Detail*.xlsx')
    latest_file = max(list_of_files, key=os.path.getctime)
    return latest_file

def IsLoginSuccessful(driver):
    username = input("Username: ")
    password = getpass("Password: ")
    
    #Login
    driver.find_element_by_id("sawlogonuser").send_keys(username)
    driver.find_element_by_id("sawlogonpwd").send_keys(password)
    driver.find_element_by_id("sawlogonpwd").send_keys(Keys.RETURN)
    driver.implicitly_wait(30)
    
    try:
        driver.find_element_by_id("sawlogonuser")
    except NoSuchElementException:
        return True
    return False

def Login(driver):
    loggedIn = False
    while not loggedIn:
        loggedIn = IsLoginSuccessful(driver)
        print("User authenticated: ", loggedIn)# Copy worksheet data to pivot sheet

def ExportTimecards(driver):
    print("Ready to extract timecards")
    driver.get("http://coloeabi03.sapient.com:9704/analytics/saw.dll?Dashboard&PortalPath=/shared/People Management/_portal/Utilization and Time Tracking")
    time.sleep(300)
    startTime = time.time()
    driver.execute_script("Download(\'saw.dll?Go&ViewID=d%3adashboard%7ep%3aldqp07kkmo07g1e0%7er%3aksbrn8ffenl8c56p&Action=Download&SearchID=jsft5oqkehr1h55pokje3s40vq&Style=sapient&PortalPath=%2fshared%2fPeople%20Management%2f_portal%2fUtilization%20and%20Time%20Tracking&Page=Timecard%20Detail&ViewState=62eqg9l029ueaedocjv58shmga&ItemName=Timecard%20Detail&path=%2fshared%2fSapient%20Reports%2fPeople%20Management%2fUtilization%20%26%20Time%20Tracking%20Dashboard%2fTimecard%20Detail&Format=excel2007&Extension=.xlsx\');")
    time.sleep(200)
    endTime = time.time()
    latestFile = getLatestPLfilepath('')
    creationTime = os.path.getctime(latestFile)
    
    if creationTime >= startTime and creationTime <= endTime:
        print("Timecards downloaded successfully")

def ExportExpenses(driver):
    print("Ready to extract expenses")
    
    #region Set query parameters
    #SetExpenseQuery(driver)
    #endregion
    time.sleep(20)

    #Generate report - Click on Apply button
    driver.find_element_by_id("gobtn").click()
    # time.sleep(300)

    # startTime = time.time()
    # driver.execute_script("Download(\'saw.dll?Go&ViewID=d%3adashboard%7ep%3ab72hrkp5man0pqek%7er%3a233omald1f5ace6g&Action=Download&SearchID=oa7qo6s9kbmmeaiopdsaoik2pe&Style=sapient&ViewState=ntihbuucj42ouircngp0brfo4i&ItemName=Profit%20and%20Loss%20Transaction%20Details&path=%2fshared%2fSapient%20Reports%2fFinancials%2fProfit%20and%20Loss%20-%20Preferred%20Currency%2fProfit%20and%20Loss%20Detail%20Level%2fProfit%20and%20Loss%20Transaction%20Details&Format=excel2007&Extension=.xlsx\');")
    # time.sleep(60)
    # endTime = time.time()
    # latestFile = getLatestPLfilepath('e')
    # creationTime = os.path.getctime(latestFile)
    
    # if creationTime >= startTime and creationTime <= endTime:
    #     print("Expenses downloaded successfully")

def SetExpenseQuery(driver):
    projNum = '190064;189623;189625;189626;189628;189629;189630;189379;189603;189560;189601'
    fiscalYr = '2018'
    fiscalQtr = 'NULL;2018 Q 1;2018 Q 2;2018 Q 3;2018 Q 4'
    billFlag = 'Y;N'
    TransType = 'Expenses'
    currency = 'INR'
    
    ids = driver.find_elements_by_css_selector("input[type='text']")
    
    driver.find_element_by_id(ids[2]).send_keys(projNum)
    driver.find_element_by_id(ids[4]).send_keys(fiscalYr)
    driver.find_element_by_id(ids[5]).send_keys(fiscalQtr)
    driver.find_element_by_id(ids[9]).send_keys(billFlag)
    driver.find_element_by_id(ids[10]).send_keys(TransType)
    driver.find_element_by_id(ids[11]).send_keys('')
    driver.find_element_by_id(ids[12]).send_keys(currency)

def MoveExpensesData():
    print('Starting copy operation for Expenses')
    
    src = getLatestPLfilepath('e')
    print('Source: ', src)
    pivotsheet = 'D:\\Box\\Box Sync\\Helios\\Helios Projects Repository\\189630 - Renault- Tech Initiatives\\Project Management\\Financials\\Profit and Loss Transaction Details.xlsx'
    print('Destination: ', pivotsheet)
    
    wb1 = xl.load_workbook(filename=src)
    ws1 = wb1.worksheets[0]
    print('Source Sheet Name: ', ws1.title)

    wb2 = xl.load_workbook(filename=pivotsheet)
    
    if 'Expenses' in wb2.sheetnames:
        del wb2['Expenses']
        print('Data sheet removed from destination')
    
    ws2 = wb2.create_sheet('Expenses')
    print('New Data sheet created in destination.')
    print('Starting copy operation')
    
    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(pivotsheet)
    print('Expenses - Copy complete')

def MoveTimesheetData():
    print('Starting copy operation for Timesheets')
    src = getLatestPLfilepath('')
    print('Source: ', src)
    
    # Rename pivot sheet
    pivotsheet = 'D:\\Box\\Box Sync\\Helios\\Helios Projects Repository\\189630 - Renault- Tech Initiatives\\Project Management\\Financials\\Profit and Loss Transaction Details.xlsx'
    print('Destination: ', pivotsheet)
    
    #Get Source sheet name
    wb1 = xl.load_workbook(filename=src)
    ws1 = wb1.worksheets[0]
    print('Source Sheet Name: ', ws1.title)

    #Check if data already exists in destination
    #Delete if existing
    wb2 = xl.load_workbook(filename=pivotsheet)
    
    if 'Timecard Detail' in wb2.sheetnames:
        del wb2['Timecard Detail']
        print('Data sheet removed from destination')
    
    #Create new data sheet for Timecards
    ws2 = wb2.create_sheet('Timecard Detail')
    print('New Data sheet created in destination. Starting copy operation')
    
    #Copy all cells to destination
    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value
            
    #Correct AH4 (Total Hours - Actual) row number
    ws2.cell(4,34).value = ws2.cell(3,34).value
    ws2.cell(3,34).value = ''
    print("Corrected Total Hours - Actual position")
    # xy = coordinate_from_string('A4') # returns ('A',4)
    #         col = column_index_from_string(xy[0]) # returns 1
    #         row = xy[1]
    #         #if row==
            
    wb2.save(pivotsheet)
    print('Timesheets - Copy complete')

def DownloadExcels():
    #region Open Chrome and Login    
    driver = webdriver.Chrome("C:\\Softwares\\chromedriver_win32\\chromedriver.exe")
    driver.get("http://coloeabi03.sapient.com:9704/analytics/saw.dll?Dashboard&PortalPath=%2Fshared%2FFinancials%2F_portal%2FProfit%20and%20Loss%20-%20Preferred%20Currency&page=PL2005%20-%20Profit%20and%20Loss%20Detail")
    driver.maximize_window()
    # driver.implicitly_wait(20)
    #endregion
    
    #region Login
    Login(driver)
    #endregion

    #region Export expenses - P&L report
    ExportExpenses(driver)
    #endregion
    
    #region Export Timetracking Report
    ExportTimecards(driver)
    # #endregion
    
    driver.quit()



